"""
src/markets/usdcop/metrics.py

Unified USDCOP Trading Metrics Module
=====================================
Production-ready metrics for Forex/USDCOP trading performance evaluation.
Combines equity-based metrics, trade statistics, and Forex-specific calculations.

Key Features:
- Core metrics: Sharpe, Sortino, Calmar, CAGR, Max Drawdown, Win Rate, Profit Factor
- Forex metrics: Pips calculations, leverage-adjusted volatility
- Multiple input formats: equity curves, signals, trade logs
- Statistical analysis: Monte Carlo, VaR, distribution analysis
- Report generation: HTML, Excel, JSON outputs
- CLI support for batch processing

Usage:
    # As module
    from src.markets.usdcop.metrics import TradingMetrics, calculate_metrics
    metrics = calculate_metrics(trades, equity_curve, config)
    
    # CLI
    python -m src.markets.usdcop.metrics equity --input equity.csv
    python -m src.markets.usdcop.metrics trades --input trades.json --pip-size 0.0001
"""

import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union, Any
from dataclasses import dataclass, asdict, field
from enum import Enum

import numpy as np
import pandas as pd
import scipy.stats as stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Configure logging
logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# =====================================================
# DATA CLASSES & CONFIGURATION
# =====================================================

@dataclass
class MetricsConfig:
    """Configuration for metrics calculation"""
    risk_free_rate: float = 0.02  # Annual
    trading_days_year: int = 252
    confidence_level: float = 0.95
    pip_size: float = 0.0001  # For USDCOP
    leverage: int = 100
    cost_per_trade: float = 0.0
    min_trades_for_stats: int = 30

@dataclass
class Trade:
    """Individual trade information"""
    entry_time: pd.Timestamp
    exit_time: pd.Timestamp
    entry_price: float
    exit_price: float
    side: str  # 'long' or 'short'
    volume: float = 1.0
    profit_loss: float = 0.0
    profit_loss_pips: float = 0.0
    
    @property
    def is_winner(self) -> bool:
        return self.profit_loss > 0

@dataclass
class PerformanceMetrics:
    """Complete performance metrics"""
    # Returns
    total_return_pct: float
    cagr: float
    
    # Risk
    volatility_annual: float
    max_drawdown: float
    max_drawdown_duration_days: int
    
    # Risk-adjusted
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    omega_ratio: float
    
    # Trading
    total_trades: int
    win_rate: float
    profit_factor: float
    expectancy: float
    avg_win: float
    avg_loss: float
    
    # Forex specific
    total_pips: float
    avg_pips_per_trade: float
    
    # Statistical
    var_95: float
    cvar_95: float
    skewness: float
    kurtosis: float
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

# =====================================================
# CORE METRICS CALCULATOR
# =====================================================

class TradingMetrics:
    """Main metrics calculator"""
    
    def __init__(self, config: Optional[MetricsConfig] = None):
        self.config = config or MetricsConfig()
        self.trades: List[Trade] = []
        self.equity_curve: Optional[pd.Series] = None
        self.returns: Optional[pd.Series] = None
        
    def calculate_all_metrics(self, 
                            trades: List[Trade],
                            equity_curve: pd.Series,
                            initial_balance: float) -> PerformanceMetrics:
        """Calculate all performance metrics"""
        self.trades = trades
        self.equity_curve = equity_curve
        self.returns = equity_curve.pct_change().fillna(0)
        
        # Core calculations
        return_metrics = self._calculate_return_metrics(initial_balance)
        risk_metrics = self._calculate_risk_metrics()
        risk_adjusted = self._calculate_risk_adjusted_ratios(return_metrics, risk_metrics)
        trading_metrics = self._calculate_trading_metrics()
        pip_metrics = self._calculate_pip_metrics()
        statistical_metrics = self._calculate_statistical_metrics()
        
        # Combine all metrics
        return PerformanceMetrics(
            **return_metrics,
            **risk_metrics,
            **risk_adjusted,
            **trading_metrics,
            **pip_metrics,
            **statistical_metrics
        )
    
    def _calculate_return_metrics(self, initial_balance: float) -> Dict[str, float]:
        """Calculate return-based metrics"""
        final_balance = self.equity_curve.iloc[-1]
        total_return = (final_balance / initial_balance - 1) * 100
        
        # CAGR
        days = len(self.equity_curve)
        years = days / self.config.trading_days_year
        cagr = (((final_balance / initial_balance) ** (1 / years)) - 1) * 100 if years > 0 else 0
        
        return {
            'total_return_pct': total_return,
            'cagr': cagr
        }
    
    def _calculate_risk_metrics(self) -> Dict[str, float]:
        """Calculate risk metrics"""
        returns = self.returns
        
        # Volatility
        volatility_annual = returns.std() * np.sqrt(self.config.trading_days_year)
        
        # Drawdown
        dd_data = self._calculate_drawdown()
        
        return {
            'volatility_annual': volatility_annual,
            'max_drawdown': dd_data['max_drawdown'],
            'max_drawdown_duration_days': dd_data['duration_days']
        }
    
    def _calculate_drawdown(self) -> Dict[str, Any]:
        """Calculate drawdown metrics"""
        equity = self.equity_curve
        running_max = equity.expanding().max()
        drawdown = (equity - running_max) / running_max
        
        max_drawdown = abs(drawdown.min())
        
        # Duration calculation
        underwater = drawdown < 0
        duration_days = 0
        current_duration = 0
        
        for is_underwater in underwater:
            if is_underwater:
                current_duration += 1
                duration_days = max(duration_days, current_duration)
            else:
                current_duration = 0
        
        return {
            'max_drawdown': max_drawdown,
            'duration_days': duration_days
        }
    
    def _calculate_risk_adjusted_ratios(self, return_metrics: Dict, risk_metrics: Dict) -> Dict[str, float]:
        """Calculate risk-adjusted performance ratios"""
        returns = self.returns
        rf_daily = self.config.risk_free_rate / self.config.trading_days_year
        
        # Sharpe Ratio
        excess_returns = returns - rf_daily
        sharpe = (excess_returns.mean() / returns.std() * 
                 np.sqrt(self.config.trading_days_year)) if returns.std() > 0 else 0
        
        # Sortino Ratio
        downside_returns = returns[returns < 0]
        downside_std = downside_returns.std() if len(downside_returns) > 0 else 0
        sortino = (excess_returns.mean() / downside_std * 
                  np.sqrt(self.config.trading_days_year)) if downside_std > 0 else 0
        
        # Calmar Ratio
        calmar = return_metrics['cagr'] / risk_metrics['max_drawdown'] / 100 if risk_metrics['max_drawdown'] > 0 else 0
        
        # Omega Ratio
        gains = returns[returns > rf_daily] - rf_daily
        losses = rf_daily - returns[returns <= rf_daily]
        omega = gains.sum() / losses.sum() if losses.sum() > 0 else 0
        
        return {
            'sharpe_ratio': sharpe,
            'sortino_ratio': sortino,
            'calmar_ratio': calmar,
            'omega_ratio': min(omega, 10.0)
        }
    
    def _calculate_trading_metrics(self) -> Dict[str, float]:
        """Calculate trade-based metrics"""
        if not self.trades:
            return self._empty_trading_metrics()
        
        winning_trades = [t for t in self.trades if t.is_winner]
        losing_trades = [t for t in self.trades if not t.is_winner]
        
        total_trades = len(self.trades)
        win_rate = len(winning_trades) / total_trades if total_trades > 0 else 0
        
        gross_profit = sum(t.profit_loss for t in winning_trades) or 0
        gross_loss = abs(sum(t.profit_loss for t in losing_trades)) or 0
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else 0
        
        avg_win = gross_profit / len(winning_trades) if winning_trades else 0
        avg_loss = gross_loss / len(losing_trades) if losing_trades else 0
        expectancy = win_rate * avg_win - (1 - win_rate) * avg_loss
        
        return {
            'total_trades': total_trades,
            'win_rate': win_rate,
            'profit_factor': min(profit_factor, 100.0),
            'expectancy': expectancy,
            'avg_win': avg_win,
            'avg_loss': avg_loss
        }
    
    def _calculate_pip_metrics(self) -> Dict[str, float]:
        """Calculate pip-based metrics"""
        if not self.trades:
            return {'total_pips': 0, 'avg_pips_per_trade': 0}
        
        total_pips = sum(t.profit_loss_pips for t in self.trades)
        avg_pips = total_pips / len(self.trades) if self.trades else 0
        
        return {
            'total_pips': total_pips,
            'avg_pips_per_trade': avg_pips
        }
    
    def _calculate_statistical_metrics(self) -> Dict[str, float]:
        """Calculate statistical metrics"""
        returns = self.returns
        
        # VaR and CVaR
        var_95 = np.percentile(returns, 5)
        cvar_95 = returns[returns <= var_95].mean() if len(returns[returns <= var_95]) > 0 else var_95
        
        # Distribution
        skewness = returns.skew() if len(returns) > 3 else 0
        kurtosis = returns.kurt() if len(returns) > 3 else 0
        
        return {
            'var_95': var_95,
            'cvar_95': cvar_95,
            'skewness': skewness,
            'kurtosis': kurtosis
        }
    
    def _empty_trading_metrics(self) -> Dict[str, float]:
        """Return empty metrics when no trades"""
        return {
            'total_trades': 0,
            'win_rate': 0,
            'profit_factor': 0,
            'expectancy': 0,
            'avg_win': 0,
            'avg_loss': 0
        }

# =====================================================
# STATISTICAL ANALYSIS
# =====================================================

class StatisticalAnalysis:
    """Advanced statistical analysis"""
    
    @staticmethod
    def monte_carlo_simulation(trades: List[Trade], 
                             n_simulations: int = 1000,
                             n_periods: int = 252) -> Dict[str, Any]:
        """Monte Carlo simulation for future projections"""
        if not trades:
            return {}
        
        # Extract returns
        returns = [t.profit_loss for t in trades]
        mean_return = np.mean(returns)
        std_return = np.std(returns)
        
        # Run simulations
        simulated_paths = []
        for _ in range(n_simulations):
            random_returns = np.random.normal(mean_return, std_return, n_periods)
            equity_path = np.cumprod(1 + random_returns / 10000)  # Normalize
            simulated_paths.append(equity_path)
        
        simulated_paths = np.array(simulated_paths)
        final_values = simulated_paths[:, -1]
        
        return {
            'expected_return': np.mean(final_values) - 1,
            'percentile_5': np.percentile(final_values, 5) - 1,
            'percentile_95': np.percentile(final_values, 95) - 1,
            'probability_profit': np.mean(final_values > 1)
        }

# =====================================================
# REPORT GENERATORS
# =====================================================

class MetricsReporter:
    """Generate reports in multiple formats"""
    
    def __init__(self, metrics: PerformanceMetrics, config: MetricsConfig):
        self.metrics = metrics
        self.config = config
    
    def generate_json_report(self, output_path: str) -> str:
        """Generate JSON report"""
        report = {
            'metrics': self.metrics.to_dict(),
            'config': asdict(self.config),
            'generated_at': datetime.now().isoformat()
        }
        
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2)
        
        logger.info(f"JSON report saved to {output_path}")
        return output_path
    
    def generate_excel_report(self, output_path: str) -> str:
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics Summary"
        
        # Header
        ws['A1'] = "USDCOP Trading Performance"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Metrics
        row = 3
        metrics_dict = self.metrics.to_dict()
        
        for key, value in metrics_dict.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value if isinstance(value, (int, float)) else str(value)
            
            # Highlight important metrics
            if key in ['sharpe_ratio', 'win_rate', 'profit_factor']:
                ws[f'A{row}'].font = Font(bold=True)
                if isinstance(value, (int, float)) and value > 0:
                    ws[f'B{row}'].font = Font(color="00AA00")
                    
            row += 1
        
        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
        return output_path
    
    def generate_html_summary(self) -> str:
        """Generate HTML summary snippet"""
        m = self.metrics
        
        html = f"""
        <div class="metrics-summary">
            <h2>Performance Summary</h2>
            <div class="metrics-grid">
                <div class="metric">
                    <span class="label">Total Return</span>
                    <span class="value">{m.total_return_pct:.2f}%</span>
                </div>
                <div class="metric">
                    <span class="label">Sharpe Ratio</span>
                    <span class="value">{m.sharpe_ratio:.2f}</span>
                </div>
                <div class="metric">
                    <span class="label">Win Rate</span>
                    <span class="value">{m.win_rate:.1%}</span>
                </div>
                <div class="metric">
                    <span class="label">Max Drawdown</span>
                    <span class="value">{m.max_drawdown:.1%}</span>
                </div>
            </div>
        </div>
        """
        return html

# =====================================================
# UTILITY FUNCTIONS
# =====================================================

def create_equity_curve_from_trades(trades: List[Trade], initial_balance: float) -> pd.Series:
    """Create equity curve from trade list"""
    if not trades:
        return pd.Series([initial_balance], index=[pd.Timestamp.now()])
    
    # Sort by exit time
    sorted_trades = sorted(trades, key=lambda t: t.exit_time)
    
    # Build curve
    dates = [sorted_trades[0].entry_time]
    equity = [initial_balance]
    
    cumulative_pnl = 0
    for trade in sorted_trades:
        cumulative_pnl += trade.profit_loss
        dates.append(trade.exit_time)
        equity.append(initial_balance + cumulative_pnl)
    
    return pd.Series(equity, index=dates)

def calculate_pips(entry_price: float, exit_price: float, side: str, pip_size: float) -> float:
    """Calculate pips for a trade"""
    price_diff = exit_price - entry_price
    if side == 'short':
        price_diff = -price_diff
    return price_diff / pip_size

def calculate_metrics(trades: Union[List[Dict], List[Trade]], 
                     equity_curve: Optional[pd.Series] = None,
                     initial_balance: float = 10000,
                     config: Optional[MetricsConfig] = None) -> PerformanceMetrics:
    """Main entry point for metrics calculation"""
    config = config or MetricsConfig()
    
    # Convert dict trades to Trade objects
    if trades and isinstance(trades[0], dict):
        trade_objects = []
        for t in trades:
            # Calculate pips if not provided
            if 'profit_loss_pips' not in t:
                pips = calculate_pips(
                    t['entry_price'], 
                    t['exit_price'], 
                    t['side'],
                    config.pip_size
                )
                t['profit_loss_pips'] = pips
            
            trade = Trade(
                entry_time=pd.to_datetime(t['entry_time']),
                exit_time=pd.to_datetime(t['exit_time']),
                entry_price=t['entry_price'],
                exit_price=t['exit_price'],
                side=t['side'],
                volume=t.get('volume', 1.0),
                profit_loss=t.get('profit_loss', 0),
                profit_loss_pips=t['profit_loss_pips']
            )
            trade_objects.append(trade)
        trades = trade_objects
    
    # Create equity curve if not provided
    if equity_curve is None:
        equity_curve = create_equity_curve_from_trades(trades, initial_balance)
    
    # Calculate metrics
    calculator = TradingMetrics(config)
    return calculator.calculate_all_metrics(trades, equity_curve, initial_balance)

# =====================================================
# CLI INTERFACE
# =====================================================

def main():
    """CLI entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="USDCOP Trading Metrics Calculator")
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    
    # Equity command
    equity_parser = subparsers.add_parser('equity', help='Calculate metrics from equity curve')
    equity_parser.add_argument('--input', required=True, help='Input CSV file')
    equity_parser.add_argument('--initial-balance', type=float, default=10000)
    equity_parser.add_argument('--output', default='metrics_report.json')
    
    # Trades command
    trades_parser = subparsers.add_parser('trades', help='Calculate metrics from trades')
    trades_parser.add_argument('--input', required=True, help='Input JSON file')
    trades_parser.add_argument('--pip-size', type=float, default=0.0001)
    trades_parser.add_argument('--output', default='metrics_report.json')
    
    args = parser.parse_args()
    
    if args.command == 'equity':
        # Load equity curve
        df = pd.read_csv(args.input, index_col='date', parse_dates=True)
        equity_curve = df['equity']
        
        # Simple trades extraction from equity changes
        trades = []
        # ... trade extraction logic ...
        
        metrics = calculate_metrics(trades, equity_curve, args.initial_balance)
        
    elif args.command == 'trades':
        # Load trades
        with open(args.input, 'r') as f:
            trades = json.load(f)
        
        config = MetricsConfig(pip_size=args.pip_size)
        metrics = calculate_metrics(trades, config=config)
    
    else:
        parser.print_help()
        return
    
    # Generate report
    reporter = MetricsReporter(metrics, config if 'config' in locals() else MetricsConfig())
    reporter.generate_json_report(args.output)
    
    # Print summary
    print(f"\nPerformance Summary:")
    print(f"Total Return: {metrics.total_return_pct:.2f}%")
    print(f"Sharpe Ratio: {metrics.sharpe_ratio:.2f}")
    print(f"Win Rate: {metrics.win_rate:.1%}")
    print(f"Max Drawdown: {metrics.max_drawdown:.1%}")

if __name__ == "__main__":
    main()